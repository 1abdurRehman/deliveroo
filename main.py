import csv
import os
from time import sleep
from datetime import datetime

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from get_mail import get_mail_out

# urls
PAYMENT_URL = 'https://deliveroo.co.uk/checkout/payment'
BASE_DIR = os.getcwd()



options = Options()
options.add_argument("--start-maximized")
# options.add_argument("--headless")
options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.3770.100 Safari/537.36')
options.add_experimental_option("prefs", {
        "profile.managed_default_content_settings.images": 2,
    })

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()),  options=options)

wbwait = WebDriverWait(driver, 10)

def convert_exp_date(date):
    formatted_date_string = date.strftime("%m/%y")
    return str(formatted_date_string)

def click_btn(ptrn, by="ID"):
    if by == "ID":
        button = wbwait.until(
                EC.presence_of_element_located((By.ID, ptrn))
            )
    else:
        button = wbwait.until(
                EC.presence_of_element_located((By.XPATH, ptrn))
            )
    button.click()

def continue_to_payment_page(magic_link):
    driver.get(magic_link)
    sleep(3)
    driver.get(PAYMENT_URL)

def continue_with_email():
    click_btn('continue-with-email')
    form_input = wbwait.until(
            EC.presence_of_element_located((By.ID, 'email-address'))
        )
    form_input.send_keys("abd.reh980@outlook.com")
    # affter filling click continue
    click_btn('//button/span[text()="Continue"]', by='XPATH')

    # after that it will ask for method to login
    click_btn('magic-link-btn')
    # get_magic_link_from_outlook
    sleep(10) #wait so that magic_email is received
    magic_link = str(get_mail_out())
    continue_to_payment_page(magic_link)

def checking_required_items():
    ele = wbwait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "MenuItemModifiers")]')))
    print(ele)
    if ele:
        req_items = ele.find_elements(By.XPATH, '//p[text()="Required"]/..')
        if not req_items:
            return False
        for item in req_items:
            item_id = item.get_attribute('id')
            radio_btn = item.find_element(By.XPATH, f'//div[@id="{item_id}"]//button//input[@type="radio"]')
            print('radio_btn name', radio_btn.get_attribute('name'))
            driver.execute_script("arguments[0].click();", radio_btn)
            print("button_clicked")
            sleep(2)

def get_file_rows(input_file):
    wb = load_workbook(input_file)
    # You can access a specific sheet by its name
    sheet = wb[wb.sheetnames[0]]
    # You can access the cell values in the sheet using the cell method
    print(sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value)
    return wb, sheet, sheet.max_row

def switch_to_iframe(name):
    iframes = {'cardnumber': 'stripe-card-number-input', 'exp-date': 'stripe-card-expiry-input', 'cvc': 'stripe-card-cvv-input', 'postal': 'stripe-postcode-input'}
    id = iframes[f'{name}']
    iframe = wbwait.until(
            EC.presence_of_element_located((By.XPATH, f'//div[@id="{id}"]//iframe'))
        )
    return driver.switch_to.frame(iframe)

def check_validity():
    """
    This function will check validity of card True if valid else False
    """
    try:
        invalid_card = wbwait.until(
            EC.presence_of_element_located((By.XPATH, '//h4[text()="Invalid Card"]'))
        )
        click_btn('//button/span[text()="Close"]', by='XPATH')
    except TimeoutException:
        invalid_card = False

    # case 1
    if invalid_card:
        return False
    # case 2
    if not change_card():
        return False
    return True

def change_card():
    try:
        change_btn = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//button/span[text()="Change"]'))
        )
        if change_btn:
            change_btn.click()
            delete_btn = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//button/span[text()="Delete"]'))
            )
            delete_btn.click()
            return True
    except TimeoutException:
        return False

def fill_card_info(info):
    change_card()
    click_btn('//label[@for="payment_method_new_card"]/..', by='XPATH')
    for name, value in info.items():
        print(f"filling {name} -> {value}")
        switch_to_iframe(name)
        input = wbwait.until(
            EC.presence_of_element_located((By.XPATH, f'//input[@name="{name}"]'))
        )
        input.clear()
        input.send_keys(value)
        sleep(1)
        driver.switch_to.default_content() # switch back
    click_btn('//button/span[text()="Add a card"]', by="XPATH")
    sleep(10) #necessary wait
    return check_validity()


def main():
    driver.get("https://deliveroo.co.uk/menu/London/fitzrovia/eggslut-fitzrovia/?day=today&geohash=gcpvh6vbtj76&time=ASAP&item-id=987074422")
    # driver.get("https://deliveroo.co.uk/menu/London/marylebone/10556-kfc-baker-street/?day=today&geohash=gcpvj0e56cwp&time=ASAP&item-id=996998500")
    # driver.get("https://deliveroo.co.uk/menu/London/strand/mcdonalds-0007-strand/?day=today&geohash=gcpvj0e56cwp&time=ASAP&item-id=1281550169")

    # driver.get("https://deliveroo.co.uk/menu/London/leicester-square/shake-shack-leicester-square/?day=today&geohash=gcpvj0e56cwp&time=ASAP&item-id=1239364389")

    #First stept to go for add button
    click_btn("onetrust-accept-btn-handler")
    checking_required_items()
    sleep(2)
    click_btn('//button/span[contains(text(), "Add for")]', by="XPATH")

    # once added go for checkout
    sleep(5)
    click_btn('//button/span[text()="Go to checkout"]', by='XPATH')
    sleep(2)
    click_btn('//button/span[text()="Continue to checkout"]', by="XPATH")

    # sign in process starts from this step
    continue_with_email()
    sleep(5)

    # now bot is on payment page
    print(driver.current_url)

    # EXCEL WORK
    valid_cards = []
    input_file = os.path.join(BASE_DIR, f"Checker_Input.xlsx")
    wb, sheet, rows_count = get_file_rows(input_file)
    for row in range(2, rows_count+1):
        card, exp, cvv, postal_c = sheet[f'A{row}'].value, sheet[f'B{row}'].value, sheet[f'C{row}'].value, sheet[f'D{row}'].value
        if card not in valid_cards:
            info = {'cardnumber': card, 'exp-date': convert_exp_date(exp), 'cvc': cvv, 'postal': postal_c}
            print('Processing:', info)
            valid_flag = fill_card_info(info)
            if valid_flag:
                sheet[f'E{row}'] = "True"
                valid_cards.append(card)
            else:
                sheet[f'E{row}'] = "False"
        print('\n')
        sleep(5)
    print(valid_cards)
    wb.save(input_file)
    wb.close()
        

main()